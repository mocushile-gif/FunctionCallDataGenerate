#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
提取表格摘要
"""

import time

start_time = time.time()

from collections import defaultdict
import csv
import datetime
import functools
from itertools import islice
import json
import math
import os
from pathlib import Path
import random
import re
import signal
import threading
import io
import traceback
import sys

import pandas as pd
import xlrd

import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from dotenv import load_dotenv
load_dotenv()

def __0_prog():
    """ 工程基础组件 """


def safe_div(a, b):
    """ 安全除法，避免除数为0的情况

    :param a: 被除数
    :param b: 除数
    :return: a/b，如果b为0，返回0
    """
    if b == 0:
        return a / sys.float_info.epsilon
    else:
        return a / b


class Timeout:
    """ 对函数等待执行的功能，限制运行时间

    【实现思路】
    1、最简单的方式是用signal.SIGALRM实现（包括三方库timeout-decorator也有这个局限）
        https://stackoverflow.com/questions/2281850/timeout-function-if-it-takes-too-long-to-finish
        但是这个不支持windows系统~~
    2、那windows和linux通用的做法，就是把原执行函数变成一个子线程来运行
        https://stackoverflow.com/questions/21827874/timeout-a-function-windows
        但是，又在onenote的win32com发现，有些功能没办法丢到子线程里，会出问题
        而且使用子线程，也没法做出支持with上下文语法的功能了
    3、于是就有了当前我自己搞出的一套机制
        是用一个Timer计时器子线程计时，当timeout超时，使用信号机制给主线程抛出一个异常
            ① 注意，不能子线程直接抛出异常，这样影响不了主线程
            ② 也不能直接抛出错误signal，这样会强制直接中断程序。应该抛出TimeoutError，让后续程序进行超时逻辑的处理
            ③ 这里是让子线程抛出信号，主线程收到信号后，再抛出TimeoutError

    注意：这个函数似乎不支持多线程
    """

    def __init__(self, seconds):
        self.seconds = seconds
        self.alarm = None

    def __call__(self, func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            # 1 如果超时，主线程收到信号会执行的功能
            def overtime(signum, frame):
                raise TimeoutError(f'function [{func.__name__}] timeout [{self.seconds} seconds] exceeded!')

            signal.signal(signal.SIGABRT, overtime)

            # 2 开一个子线程计时器，超时的时候发送信号
            def send_signal():
                signal.raise_signal(signal.SIGABRT)

            alarm = threading.Timer(self.seconds, send_signal)
            alarm.start()

            # 3 执行主线程功能
            res = func(*args, **kwargs)
            alarm.cancel()  # 正常执行完则关闭计时器

            return res

        return wrapper

    def __enter__(self):
        def overtime(signum, frame):
            raise TimeoutError(f'with 上下文代码块运行超时 > [{self.seconds} 秒]')

        signal.signal(signal.SIGABRT, overtime)

        def send_signal():
            signal.raise_signal(signal.SIGABRT)

        self.alarm = threading.Timer(self.seconds, send_signal)
        self.alarm.start()

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.alarm.cancel()


def run_once(distinct_mode=0, *, limit=1):
    """ 装饰器，装饰的函数在一次程序里其实只会运行一次
    :param int|str distinct_mode:
        0，默认False，不区分输入的参数值（包括cls、self），强制装饰的函数只运行一次
        'str'，设为True或1时，仅以字符串化的差异判断是否是重复调用，参数不同，会判断为不同的调用，每种调用限制最多执行limit次
        'id,str'，在'str'的基础上，第一个参数使用id代替。一般用于类方法、对象方法的装饰。
            不考虑类、对象本身的内容改变，只要还是这个类或对象，视为重复调用。
        'ignore,str'，首参数忽略，第2个开始的参数使用str格式化
            用于父类某个方法，但是子类继承传入cls，原本id不同会重复执行
            使用该模式，首参数会ignore忽略，只比较第2个开始之后的参数
        func等callable类型的对象也行，是使用run_once装饰器的简化写法
    :param limit: 默认只会执行一次，该参数可以提高限定的执行次数，一般用不到，用于兼容旧的 limit_call_number 装饰器
    returns: 返回decorator
    """
    if callable(distinct_mode):
        # @run_once，没写括号的时候去装饰一个函数，distinct_mode传入的是一个函数func
        # 使用run_once本身的默认值
        return run_once()(distinct_mode)

    def get_tag(args, kwargs):
        if not distinct_mode:
            ls = tuple()
        elif distinct_mode == 'str':
            ls = (str(args), str(kwargs))
        elif distinct_mode == 'id,str':
            ls = (id(args[0]), str(args[1:]), str(kwargs))
        elif distinct_mode == 'ignore,str':
            ls = (str(args[1:]), str(kwargs))
        else:
            raise ValueError
        return ls

    def decorator(func):
        counter = {}  # 映射到一个[cnt, last_result]

        def wrapper(*args, **kwargs):
            tag = get_tag(args, kwargs)
            if tag not in counter:
                counter[tag] = [0, None]
            x = counter[tag]
            if x[0] < limit:
                res = func(*args, **kwargs)
                x = counter[tag] = [x[0] + 1, res]
            return x[1]

        return wrapper

    return decorator


def inject_members(from_obj, to_obj, member_list=None, *,
                   check=False, ignore_case=False,
                   white_list=None, black_list=None):
    """ 将from_obj的方法注入到to_obj中
    文档：https://www.yuque.com/code4101/journal/uum4f0#kskSj

    一般用于类继承中，将子类from_obj的新增的成员方法，添加回父类to_obj中
        反经合道：这样看似很违反常理，父类就会莫名其妙多出一些可操作的成员方法。
            但在某些时候能保证面向对象思想的情况下，大大简化工程代码开发量。
    也可以用于模块等方法的添加

    :param from_obj: 一般是一个类用于反向继承的方法来源，但也可以是模块等任意对象。
        注意py一切皆类，一个class定义的类本质也是type类定义出的一个对象
        所以这里概念上称为obj才是准确的，反而是如果叫from_cls不太准确，虽然这个方法主要确实都是用于class类
    :param to_obj: 同from_obj，要被注入方法的对象
    :param Sequence[str] member_list: 手动指定的成员方法名，可以不指定，自动生成
    :param check: 检查重名方法
    :param ignore_case: 忽略方法的大小写情况，一般用于win32com接口
    :param Sequence[str] white_list: 白名单。无论是否重名，这里列出的方法都会被添加
    :param Sequence[str] black_list: 黑名单。这里列出的方法不会被添加

    # 把XlDocxTable的成员方法绑定到docx.table.Table里
    >> inject_members(XlDocxTable, docx.table.Table)

    """
    # 1 整理需要注入的方法清单
    dst = set(dir(to_obj))
    if ignore_case:
        dst = {x.lower() for x in dst}

    if member_list:
        src = set(member_list)
    else:
        if ignore_case:
            src = {x for x in dir(from_obj) if (x.lower() not in dst)}
        else:
            src = set(dir(from_obj)) - dst

    # 2 微调
    if white_list:
        src |= set(white_list)
    if black_list:
        src -= set(black_list)

    # 3 注入方法
    for x in src:
        if x.startswith("__") and x.endswith("__"):
            continue
        setattr(to_obj, x, getattr(from_obj, x))


def grp_chinese_char():
    return r'[\u4e00-\u9fa5，。；？（）【】、①-⑨]'


def calc_chinese_ratio(s):
    """ 计算中文字符比例

    >>> calc_chinese_ratio('abc')
    0.0
    >>> calc_chinese_ratio('abc中文')
    0.5714285714285714
    """
    s2 = re.sub(grp_chinese_char(), '', s)
    b = len(s2)
    a = 2 * (len(s) - b)
    # 一个汉字占2个字符权重
    return safe_div(a, a + b)


def shuffle_dict_keys(d):
    keys = list(d.keys())
    random.shuffle(keys)
    d = {k: d[k] for k in keys}
    return d


class DictTool:
    @classmethod
    def ior(cls, dict_, *args):
        """ 合并到第1个字典

        :return: dict_ |= (args[0] | args[1] | ... | args[-1]).

        220601周三15:45，默认已有对应key的话，值是不覆盖的，如果要覆盖，直接用update就行了，不需要这个接口
            所以把3.9的|=功能关掉
        """
        # if sys.version_info.major == 3 and sys.version_info.minor >= 9:
        #     for x in args:
        #         dict_ |= x
        # else:  # 旧版本py手动实现一个兼容功能
        for x in args:
            for k, v in x.items():
                # 220729周五21:21，又切换成dict_有的不做替换
                if k not in dict_:
                    dict_[k] = v
                # dict_[k] = v


def human_readable_number(value, base_type='K', precision=4):
    """ 数字美化输出函数

    :param float|int value: 要转换的数值
    :param int precision: 有效数字的长度
    :param str base_type: 进制类型，'K'为1000进制, 'KB'为1024进制（KiB同理）, '万'为中文万进制
    :return: 美化后的字符串
    """
    if abs(value) < 1:
        return f'{value:.{precision}g}'

    # 设置不同进制的单位和基数
    units, base = {
        'K': (['', 'K', 'M', 'G', 'T', 'P', 'E', 'Z', 'Y'], 1000),
        'KB': (['', 'KB', 'MB', 'GB', 'TB', 'PB', 'EB', 'ZB', 'YB'], 1024),
        'KiB': (['', 'KiB', 'MiB', 'GiB', 'TiB', 'PiB', 'EiB', 'ZiB', 'YiB'], 1024),
        '万': (['', '万', '亿', '万亿', '亿亿'], 10000)
    }.get(base_type, ([''], 1))  # 默认为空单位和基数1

    x, i = abs(value), 0
    while x >= base and i < len(units) - 1:
        x /= base
        i += 1

    x = f'{x:.{precision}g}'  # 四舍五入到指定精度
    prefix = '-' if value < 0 else ''  # 负数处理
    return f"{prefix}{x}{units[i]}"


def format_exception(e, mode=3):
    if mode == 1:
        # 仅获取异常类型的名称
        text = ''.join(traceback.format_exception_only(type(e), e)).strip()
    elif mode == 2:
        # 获取异常类型的名称和附加的错误信息
        text = f"{type(e).__name__}: {e}"
    elif mode == 3:
        text = ''.join(traceback.format_exception(type(e), e, e.__traceback__))
    else:
        raise ValueError
    return text


def __1_basic():
    """ 表格的组件功能 """


def excel_addr(n, m) -> str:
    r"""数字索引转excel地址索引

    :param n: 行号，可以输入字符串形式的数字
    :param m: 列号，同上可以输入str的数字
    :return:

    >>> excel_addr(2, 3)
    'C2'
    """
    return f'{get_column_letter(int(m))}{n}'


def excel_addr2(n1, m1, n2, m2) -> str:
    r""" excel_addr的扩展版，定位一个区间

    >>> excel_addr2(2, 3, 4, 4)
    'C2:D4'
    """
    return f'{get_column_letter(int(m1))}{n1}:{get_column_letter(int(m2))}{n2}'


def is_valid_excel_cell(cell):
    """ 判断输入的字符串是否是一个合法的Excel单元格地址

    :param str cell: 输入的字符串
    :return bool: 如果是合法的Excel单元格地址返回True，否则返回False
    """
    match = re.fullmatch(r'[A-Z]+[1-9][0-9]*', cell)
    return match is not None


def is_valid_excel_range(range):
    """ 判断输入的字符串是否是一个合法的Excel单元格范围

    :param str range: 输入的字符串
    :return bool: 如果是合法的Excel单元格范围返回True，否则返回False
    """
    if ':' in range:
        start, end = range.split(':')
        return (is_valid_excel_cell(start) or start.isdigit() or re.fullmatch(r'[A-Z]+', start)) and \
            (is_valid_excel_cell(end) or end.isdigit() or re.fullmatch(r'[A-Z]+', end)) and \
            start <= end
    else:
        return is_valid_excel_cell(range)


def is_valid_excel_address(address):
    """ 判断输入的字符串是否是一个合法的Excel地址定位

    :param str address: 输入的字符串
    :return bool: 如果是合法的Excel地址定位返回True，否则返回False

    注意，严格来说，'A1,A3'这种定位也是可以的，但是这个函数暂不考虑这种情况，
        如果需要，可以另外写is_valid_excel_address2
    """
    if ':' in address:
        return is_valid_excel_range(address)
    else:
        return is_valid_excel_cell(address)


@run_once('str')
def xlfmt2pyfmt_date(xl_fmt):
    """ 日期的渲染操作

    >>> xlfmt2pyfmt_date('yyyy/m/d')
    '%Y/{month}/{day}'
    >>> xlfmt2pyfmt_date('yyyy-mm-dd')
    '%Y-%m-%d'
    >>> xlfmt2pyfmt_date('yyyy年mm月dd日')
    '%Y年%m月%d日'
    >>> xlfmt2pyfmt_date('yyyy年m月d日')
    '%Y年{month}月{day}日'

    # 注意以下是并未支持的功能...会默认返回 '%Y/%-m/%-d'
    >>> xlfmt2pyfmt_date('yy-m-d')
    '%y-{month}-{day}'
    >>> xlfmt2pyfmt_date('m/d/yy')
    '{month}/{day}/%y'
    >>> xlfmt2pyfmt_date('dddd, mmmm dd, yyyy')
    '%A, %B {day}, %Y'
    >>> xlfmt2pyfmt_date('yy/mm/dd')
    '%y/%m/%d'
    >>> xlfmt2pyfmt_date('m-d-yy')
    '{month}-{day}-%y'
    """
    mappings = {
        'y': {2: '%y', 4: '%Y'},
        'm': {1: '%-m', 2: '%m', 3: '%b', 4: '%B'},
        'd': {1: '%-d', 2: '%d', 3: '%a', 4: '%A'}
    }

    m = re.search(r'(y+)(.+?)(m+)(.+?)(d+)(日?)', xl_fmt.replace('"', ''))
    if m:
        y, sep1, m, sep2, d, sep3 = m.groups()
        year_pattern = mappings['y'].get(len(y), '%Y')
        month_pattern = mappings['m'].get(len(m), '%m')
        day_pattern = mappings['d'].get(len(d), '%d')
        fmt = f'{year_pattern}{sep1}{month_pattern}{sep2}{day_pattern}{sep3}'
    else:
        fmt = '%Y/%-m/%-d'

    # 在windows下，%-m和%-d会报错，所以需要替换成{month}和{day}
    fmt = fmt.replace('%-m', '{month}')
    fmt = fmt.replace('%-d', '{day}')
    return fmt


@run_once('str')
def xlfmt2pyfmt_time(xl_fmt):
    """ 时间的渲染操作

    >>> xlfmt2pyfmt_time('h:mm:ss')
    '%I:%M:%S'
    >>> xlfmt2pyfmt_time('hh:mm:ss')
    '%H:%M:%S'
    >>> xlfmt2pyfmt_time('mm:ss')
    '%M:%S'
    >>> xlfmt2pyfmt_time('h:mm')
    '%I:%M'
    >>> xlfmt2pyfmt_time('hh:mm')
    '%H:%M'
    >>> xlfmt2pyfmt_time('m:ss')
    '%M:%S'
    >>> xlfmt2pyfmt_time('h:mm:ss AM/PM')
    '%I:%M:%S %p'
    >>> xlfmt2pyfmt_time('hh:mm:ss AM/PM')
    '%H:%M:%S %p'
    """
    xl_fmt = re.sub(r'(y+)(.+?)(m+)(.+?)(d+)(日?)', '', xl_fmt.replace('"', ''))

    components = []

    # 判断是12小时制还是24小时制
    if 'hh' in xl_fmt:
        components.append('%H')
    elif 'h' in xl_fmt:
        components.append('%I')

    # 判断是否显示分钟
    if 'mm' in xl_fmt or 'm' in xl_fmt:
        components.append('%M')

    # 判断是否显示秒钟
    if 'ss' in xl_fmt or 's' in xl_fmt:
        components.append('%S')

    # 判断是否显示AM/PM
    if 'AM/PM' in xl_fmt:
        if components:
            components[-1] += ' %p'
        else:
            components.append('%p')

    return ':'.join(components)


def xlfmt2pyfmt_datetime(xl_fmt):
    """ 主要是针对日期、时间的渲染操作

    >>> xlfmt2pyfmt_datetime('yyyy-mm-dd h:mm:ss')
    '%Y-%m-%d %I:%M:%S'
    """
    py_fmt = xlfmt2pyfmt_date(xl_fmt)
    if ':' in xl_fmt:
        py_fmt += ' ' + xlfmt2pyfmt_time(xl_fmt)
    return py_fmt


def xl_render_value(x, xl_fmt):
    """ 得到单元格简单渲染后的效果
    py里不可能对excel的所有格式进行全覆盖，只是对场景格式进行处理

    注意，遇到公式是很难计算处理的，大概率只能保持原公式显示
    因为日期用的比较多，需要时常获得真实的渲染效果，所以这里封装一个接口

    对于JSA等场景，直接使用Cell.Text获取渲染值就行，不需要这里这么复杂的实现

    >>> xl_render_value(datetime.datetime(2020, 1, 1), 'yyyy-mm-dd')
    '2020-01-01'
    """

    if isinstance(x, datetime.datetime):
        y = x.strftime(xlfmt2pyfmt_datetime(xl_fmt)).format(month=x.month, day=x.day)
    elif isinstance(x, datetime.date):
        y = x.strftime(xlfmt2pyfmt_date(xl_fmt)).format(month=x.month, day=x.day)
    elif isinstance(x, datetime.time):
        y = x.strftime(xlfmt2pyfmt_time(xl_fmt))
    elif isinstance(x, datetime.timedelta):
        y = str(x)
    elif isinstance(x, (str, int, float, bool)):  # 其他可以json化的数据类型
        y = x
    else:  # ArrayFormula、DataTableFormula等无法json化的数据，提前转str
        y = str(x)
    return y


def convert_csv_text_to_xlsx(csv_text):
    """ 将 csv 文本转换为 xlsx 文件 """
    wb = Workbook()
    sheet = wb.active

    # 使用 io.StringIO 将 csv 文本转换为可读的文件对象
    f = io.StringIO(csv_text)
    reader = csv.reader(f)
    for row_idx, row in enumerate(reader, start=1):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx).value = value

    f.close()
    return wb


def convert_csv_to_xlsx(csv_file):
    """ 将 csv 文件转换为 xlsx 文件 """
    wb = Workbook()
    sheet = wb.active

    with open(csv_file, encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx).value = value

    return wb


def convert_xls_to_xlsx(xls_file):
    """ 将 xls 文件转换为 xlsx 文件

    注意，这只是一个简化版的转换，要尽量完整的话，还是要用microsoft 365来升级xls的
    """
    # 使用 xlrd 打开 xls 文件
    xls_workbook = xlrd.open_workbook(xls_file)

    # 创建一个新的 openpyxl 工作簿
    wb = Workbook()
    sheet = wb.active

    for i in range(xls_workbook.nsheets):
        xls_sheet = xls_workbook.sheet_by_index(i)
        if i > 0:
            sheet = wb.create_sheet(xls_sheet.name)
        for row in range(xls_sheet.nrows):
            for col in range(xls_sheet.ncols):
                # 将 xlrd 单元格的数据写入 openpyxl 单元格
                sheet.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)

    return wb


def parse_range_address(address):
    """ 解析单元格范围地址。

    :param str address: 单元格范围地址，例如 'A1', 'A1:B3', '1:3', 'A:B' 等。
    :return dict: 一个包含 'left', 'top', 'right', 'bottom' 的字典。
    """
    # 初始化默认值
    left, right, top, bottom = None, None, None, None

    # 分割地址以获取开始和结束
    parts = address.split(":")
    start_cell = parts[0]
    end_cell = parts[1] if len(parts) > 1 else start_cell

    # 如果 start_cell 是行号
    if start_cell.isdigit():
        top = int(start_cell)
    else:
        # 尝试从 start_cell 提取列
        try:
            left = column_index_from_string(start_cell.rstrip('1234567890'))
            top = int(''.join(filter(str.isdigit, start_cell))) if any(
                char.isdigit() for char in start_cell) else None
        except ValueError:
            left = None

    # 如果 end_cell 是行号
    if end_cell.isdigit():
        bottom = int(end_cell)
    else:
        # 尝试从 end_cell 提取列
        try:
            right = column_index_from_string(end_cell.rstrip('1234567890'))
            bottom = int(''.join(filter(str.isdigit, end_cell))) if any(char.isdigit() for char in end_cell) else None
        except ValueError:
            right = None

    # 如果只提供了一个部分 (例如 '1', 'A')，将最大值设置为最小值
    if len(parts) == 1:
        right = left if left is not None else right
        bottom = top if top is not None else bottom

    return {"left": left, "top": top, "right": right, "bottom": bottom}


def get_addr_area(addr):
    """ 一个range描述的面积 """
    if ':' in addr:
        d = parse_range_address(addr)
        return (d['right'] - d['left'] + 1) * (d['bottom'] - d['top'] + 1)
    else:
        return 1


def build_range_address(left=None, top=None, right=None, bottom=None):
    """ 构建单元格范围地址。

    :return str: 单元格范围地址，例如 'A1', 'A1:B3', '1:3', 'A:B' 等。
    """
    start_cell = f"{get_column_letter(left) if left else ''}{top if top else ''}"
    end_cell = f"{get_column_letter(right) if right else ''}{bottom if bottom else ''}"

    # 当开始和结束单元格相同时，只返回一个单元格地址
    if start_cell == end_cell:
        return start_cell
    # 当其中一个单元格是空字符串时，只返回另一个单元格地址
    elif not start_cell or not end_cell:
        return start_cell or end_cell
    else:
        return f"{start_cell}:{end_cell}"


def combine_addresses(*addrs):
    # 初始化最小和最大行列值
    min_left, min_top, max_right, max_bottom = float('inf'), float('inf'), 0, 0

    # 遍历所有地址
    for addr in addrs:
        # 解析每个地址
        addr_dict = parse_range_address(addr)

        # 更新最小和最大行列值
        if addr_dict['left'] is not None:
            min_left = min(min_left, addr_dict['left'])
            max_right = max(max_right, addr_dict['right'] if addr_dict['right'] is not None else addr_dict['left'])
        if addr_dict['top'] is not None:
            min_top = min(min_top, addr_dict['top'])
            max_bottom = max(max_bottom, addr_dict['bottom'] if addr_dict['bottom'] is not None else addr_dict['top'])

    # 构建新的地址字符串
    new_addr = f"{get_column_letter(min_left)}{min_top}:{get_column_letter(max_right)}{max_bottom}"
    return new_addr


def is_string_type(value):
    """ 检查值是否为字符串类型，不是数值或日期类型 """
    # 首先检查日期类型
    try:
        pd.to_datetime(value, errors='raise')
        return False
    except (ValueError, TypeError, OverflowError):
        pass

    # 检查是否为浮点数类型
    try:
        float(value)
        return False
    except (ValueError, TypeError):
        return True


def __2_openpyxl_class():
    """ 对openpyxl已有类的功能的增强 """


class XlCell(openpyxl.cell.cell.Cell):  # 适用于 openpyxl.cell.cell.MergedCell，但这里不能多重继承

    def in_range(self):
        """ 判断一个单元格所在的合并单元格

        >> ws['C1'].in_range()
        <openpyxl.worksheet.cell_range.CellRange> A1:D3
        """
        ws = self.parent
        for rng in ws.merged_cells.ranges:
            if self.coordinate in rng:
                break
        else:  # 如果找不到则返回原值
            rng = self
        return rng

    def mcell(self):
        """返回“有效单元格”，即如果输入的是一个合并单元格，会返回该合并单元格左上角的单元格
        修改左上角单元格的值才是可行、有意义的

        因为跟合并单元格有关，所以 以m前缀 merge
        """
        if isinstance(self, MergedCell):
            ws = self.parent
            x, y = self.in_range().top[0]
            return ws.cell(x, y)
        else:
            return self

    def celltype(self, *, return_mode=False):
        """
        :param return_mode: 是否返回运算的中间结果信息
            主要是在type=2的情景，有时候需要使用rng变量，可以这里直接返回，避免外部重复计算
        :return: 单元格类型
            0：普通单元格
            1：合并单元格其他衍生位置
            2：合并单元格的左上角的位置

        TODO 这个函数还是可以看看能不能有更好的实现、提速
        """
        _type, status = 0, {}
        if isinstance(self, MergedCell):
            _type = 1
        elif isinstance(self.offset(1, 0), MergedCell) or isinstance(self.offset(0, 1), MergedCell):
            # 这里只能判断可能是合并单元格，具体是不是合并单元格，还要
            rng = self.in_range()
            status['rng'] = rng
            _type = 2 if hasattr(rng, 'size') else 0

        if return_mode:
            return _type, status
        else:
            return _type

    def isnone(self):
        """ 是普通单元格且值为None

        注意合并单元格的衍生单元格不为None
        """
        celltype = self.celltype()
        return celltype == 0 and self.value is None

    def down(self, count=1):
        """ 输入一个单元格，向下移动一格
        注意其跟offset的区别，如果cell是合并单元格，会跳过自身的衍生单元格

        :param count: 重复操作次数

        注意这里移动跟excel中操作也不太一样，设计的更加"原子化"，可以多配合cell.mcell功能使用。
        详见：【腾讯文档】cell移动机制说明 https://docs.qq.com/doc/DUkRUaFhlb3l4UG1P
        """

        def _func(cell):
            r, c = cell.row, cell.column
            if cell.celltype():
                rng = cell.in_range()
                r = rng.max_row
            return cell.parent.cell(r + 1, c)

        cell = self
        for _ in range(count):
            cell = _func(cell)
        return cell

    def right(self, count=1):
        def _func(cell):
            r, c = cell.row, cell.column
            if cell.celltype():
                rng = cell.in_range()
                c = rng.max_col
            return cell.parent.cell(r, c + 1)

        cell = self
        for _ in range(count):
            cell = _func(cell)
        return cell

    def up(self, count=1):
        def _func(cell):
            r, c = cell.row, cell.column
            if cell.celltype():
                rng = cell.in_range()
                r = rng.min_row
            return cell.parent.cell(max(r - 1, 1), c)

        cell = self
        for _ in range(count):
            cell = _func(cell)
        return cell

    def left(self, count=1):
        def _func(cell):
            r, c = cell.row, cell.column
            if cell.celltype():
                rng = cell.in_range()
                r = rng.min_row
            return cell.parent.cell(r, max(c - 1, 1))

        cell = self
        for _ in range(count):
            cell = _func(cell)
        return cell

    def get_number_format(self):
        """ 相比源生的接口，有做了一些细节优化 """
        fmt = self.number_format
        # openpyxl的机制，如果没有配置日期格式，读取到的是默认的'mm-dd-yy'，其实在中文场景，默认格式应该是后者
        if fmt == 'mm-dd-yy':
            return 'yyyy/m/d'  # 中文的默认日期格式
        elif fmt == 'yyyy\-mm\-dd':  # 不知道为什么会有提取到这种\的情况，先暴力替换了
            fmt = 'yyyy-mm-dd'
        return fmt

    def get_render_value(self):
        """ 得到单元格简单渲染后的效果
        py里不可能对excel的所有格式进行全覆盖，只是对场景格式进行处理

        注意，遇到公式是很难计算处理的，大概率只能保持原公式显示
        因为日期用的比较多，需要时常获得真实的渲染效果，所以这里封装一个接口
        """
        x = self.value
        xl_fmt = self.get_number_format()
        return xl_render_value(x, xl_fmt)

    def address(self):
        if isinstance(self, openpyxl.cell.cell.Cell):
            return self.coordinate
        else:  # 否则认为是合并单元格
            return str(self)  # 标准库有自带方法，可以直接转的


# 只有cell和mergecell都共同没有的成员方法，才添加进去
__members = set(dir(XlCell)) - set(dir(openpyxl.cell.cell.Cell)) - \
            set(dir(openpyxl.cell.cell.MergedCell)) - {'__dict__'}
inject_members(XlCell, openpyxl.cell.cell.Cell, __members)
inject_members(XlCell, openpyxl.cell.cell.MergedCell, __members)


class XlWorksheet(openpyxl.worksheet.worksheet.Worksheet):
    """ 扩展标准的Workshhet功能 """

    def get_raw_usedrange(self):
        raw_used_range = build_range_address(left=self.min_column, top=self.min_row,
                                             right=self.max_column, bottom=self.max_row)
        return raw_used_range

    def is_empty_row(self, row, start_col, end_col):
        if not hasattr(self, 'is_empty_row_cache'):
            self.is_empty_row_cache = {}
        key = (row, start_col, end_col)

        def is_empty_row_core():
            cur_col = start_col
            # 特地提前检查下最后一列的那个单元格
            if self.cell(row, end_col).value is not None:
                return False
            while cur_col <= end_col:
                if self.cell(row, cur_col).value is not None:
                    return False
                # 步长随着尝试的增加，也逐渐降低采样率
                n = cur_col - start_col + 1
                # 在最大值m=16384列情况下，/1000，最多检索3404个单元格，/100，最多检索569次，/50最多检索320次
                # cur_col += (n // 50) + 1
                # 再变形，加强前面权重，大大降低后面权重
                if n <= 100:
                    cur_col += 1
                else:  # 最多54次
                    cur_col += (n // 10)

            return True

        if key not in self.is_empty_row_cache:
            self.is_empty_row_cache[key] = is_empty_row_core()

        return self.is_empty_row_cache[key]

    def is_empty_column(self, col, start_row, end_row):
        if not hasattr(self, 'is_empty_column_cache'):
            self.is_empty_column_cache = {}
        key = (col, start_row, end_row)

        def is_empty_column_core():
            cur_row = start_row
            # 特地提前检查下最后一行的那个单元格
            if self.cell(end_row, col).value is not None:
                return False
            while cur_row <= end_row:
                if self.cell(cur_row, col).value is not None:
                    return False
                n = cur_row - start_row + 1
                # 在最大值n=1048576行情况下，/1000，最多检索7535个单元格，/100，最多检索987次，/50最多检索530次
                cur_row += (n // 1000) + 1
            return True

        if key not in self.is_empty_column_cache:
            self.is_empty_column_cache[key] = is_empty_column_core()

        return self.is_empty_column_cache[key]

    def find_last_non_empty_row(self, start_row, end_row, start_col, end_col, m=30):
        # 1 如果剩余行数不多（小于等于m），直接遍历这些行
        if end_row - start_row <= m:  # 这里是兼容start_row大于end_row的情况的
            for row in range(end_row, start_row - 1, -1):
                if not self.is_empty_row(row, start_col, end_col):
                    return row
            return -1  # 如果都是空的，则返回-1

        # 2 计算分割点
        intervals = [(end_row - start_row) // (m - 1) * i + start_row for i in range(m - 1)] + [end_row]

        # 3 反向遍历这些分割点，找到第一个非空行
        for i in reversed(range(len(intervals))):
            # 检查点全部都空，也不能判定全空，要检查前面半个区间
            if i == 0 or not self.is_empty_row(intervals[i], start_col, end_col):
                # 如果这是最后一个分割点，则直接返回它
                if i == m - 1:
                    return intervals[i]
                # 否则，在这个分割点和下一个（一半二分的位置）之间递归查找
                return self.find_last_non_empty_row(intervals[i],
                                                    intervals[min(i + m // 2, m - 1)],
                                                    start_col,
                                                    end_col,
                                                    m + 1)

        # 如果所有分割点都是空的，则返回-1
        return -1

    def find_last_non_empty_column(self, start_col, end_col, start_row, end_row, m=30):
        # dprint(end_col)

        # 1 如果剩余列数不多（小于等于m），直接遍历这些列
        if end_col - start_col <= m:
            for col in range(end_col, start_col - 1, -1):
                if not self.is_empty_column(col, start_row, end_row):
                    return col
            return -1  # 如果都是空的，则返回-1

        # 2 计算分割点
        intervals = [(end_col - start_col) // (m - 1) * i + start_col for i in range(m - 1)] + [end_col]

        # 3 反向遍历这些分割点，找到第一个非空列
        for i in reversed(range(len(intervals))):
            if i == 0 or not self.is_empty_column(intervals[i], start_row, end_row):
                # 如果这是最后一个分割点，则直接返回它
                if i == m - 1:
                    return intervals[i]
                # 否则，在这个分割点和下一个（一半二分的位置）之间递归查找
                return self.find_last_non_empty_column(intervals[i],
                                                       intervals[min(i + m // 2, m - 1)],
                                                       start_row,
                                                       end_row,
                                                       m + 1)
        # 如果所有分割点都是空的，则返回-1
        return -1

    def find_first_non_empty_row(self, start_row, end_row, start_col, end_col, m=30):
        # 1 如果剩余行数不多（小于等于m），直接遍历这些行
        if end_row - start_row <= m:
            for row in range(start_row, end_row + 1):
                if not self.is_empty_row(row, start_col, end_col):
                    return row
            return -1  # 如果都是空的，则返回-1

        # 2 计算分割点
        intervals = [(end_row - start_row) // (m - 1) * i + start_row for i in range(m - 1)] + [end_row]

        # 3 正向遍历这些分割点，找到第一个非空行
        for i in range(len(intervals)):
            if i == m - 1 or not self.is_empty_row(intervals[i], start_col, end_col):
                # 如果这是第一个分割点，则直接返回它
                if i == 0:
                    return intervals[i]
                # 否则，在这个分割点和前一个（一半二分的位置）之间递归查找
                return self.find_first_non_empty_row(intervals[max(i - m // 2, 0)],
                                                     intervals[i],
                                                     start_col,
                                                     end_col,
                                                     m + 1)
        # 如果所有分割点都是空的，则返回-1
        return -1

    def find_first_non_empty_column(self, start_col, end_col, start_row, end_row, m=30):
        # 1 如果剩余列数不多（小于等于m），直接遍历这些列
        if end_col - start_col <= m:
            for col in range(start_col, end_col + 1):
                if not self.is_empty_column(col, start_row, end_row):
                    return col
            return -1  # 如果都是空的，则返回-1

        # 2 计算分割点
        intervals = [(end_col - start_col) // (m - 1) * i + start_col for i in range(m - 1)] + [end_col]

        # 3 正向遍历这些分割点，找到第一个非空列
        for i in range(len(intervals)):
            if i == m - 1 or not self.is_empty_column(intervals[i], start_row, end_row):
                # 如果这是第一个分割点，则直接返回它
                if i == 0:
                    return intervals[i]
                # 否则，在这个分割点和前一个（一半二分的位置）之间递归查找
                return self.find_first_non_empty_column(intervals[max(i - m // 2, 0)],
                                                        intervals[i],
                                                        start_row,
                                                        end_row,
                                                        m + 1)
        # 如果所有分割点都是空的，则返回-1
        return -1

    def get_usedrange(self):
        """ 定位有效数据区间。

        背景：
            在Excel工作表中，经常需要确定包含数据的有效区域。这是因为工作表可能包含大量的空白区域，
            而实际数据仅占据一部分空间。有效地识别这个区域对于进一步的数据处理和分析至关重要。
            如果暴力一行行遍历，遇到XFD1048576这种覆盖全范围的表，运行肯定会超时。

        求解思路：
            为了高效地定位有效数据区间，我们采用了分割点技术和二分查找的思想。
            具体解释，以找第1~100行中最后一个非空行为例。
            可以判断第1、50、100行是否是空行，如果50、100都是空行，那空行应该在第1~50的范围里，然后再判定第25行是否为空行。
            但二分法可能不严谨，第50、100都是空行，中间也有可能有内容，比如第1~80行本来其实都有内容，只是恰好第50行空了。
            所以在二分法的基础上，还需要把1~100行等间距取m个采样点来辅助检查。

            通过调m的值的变化方法，可以在速度和精度之间做一个权衡。
            这四个定界符，最最慢的是find_last_non_empty_row，重点调这个。

        四个边界判定函数，工程上是可以整合的，但是整合后，太多if分支操作，会降低效率，这里为了运行速度，就拆分4个实现更好。
            具体实现中，还有其他一些细节考虑优化。
            比如先找最后行，再最后列，再第一行，第一列，这个顺序是有讲究的，可以减少很多不必要的遍历。
            因为数据一般是比较高和窄的，所以应该先对行做处理。以及前面出现空区域的概率小，可以后面再处理。
            而且就openpyxl而言，对列的遍历也是远慢于行的遍历的。

        :param reset_bounds: 计算出新区域后，是否重置ws的边界值
        """
        if not hasattr(self, 'usedrange_cache'):
            # 初始化边界值
            left, right, top, bottom = self.min_column, self.max_column, self.min_row, self.max_row

            # start_time = time.time()
            # 使用优化后的函数找到最下方的行和最右边的列
            bottom = self.find_last_non_empty_row(top, bottom, left, right)
            if bottom == -1:
                return 'A1'  # 空表返回A1占位
            right = self.find_last_non_empty_column(left, right, top, bottom)
            if right == -1:
                return 'A1'

            # 使用优化后的函数找到最上方的行和最左边的列
            top = self.find_first_non_empty_row(top, bottom, left, right)
            if top == -1:
                return 'A1'
            left = self.find_first_non_empty_column(left, right, top, bottom)
            if left == -1:
                return 'A1'
            # get_global_var('get_usedrange_time')[-1] += time.time() - start_time

            # 2 然后还要再扩范围（根据合并单元格情况）
            # start_time = time.time()
            top0, bottom0, left0, right0 = top, bottom, left, right
            for merged_range in self.merged_cells.ranges:
                l, t, r, b = merged_range.bounds
                if top0 <= b <= bottom0 or top0 <= t <= bottom0:
                    if left0 <= r and l < left:
                        left = l
                    if l <= right0 and r > right:
                        right = r
                if left0 <= r <= right0 or left0 <= l <= right0:
                    if top0 <= b and t < top:
                        top = t
                    if t <= bottom0 and b > bottom:
                        bottom = b
            # get_global_var('expandrange_time')[-1] += time.time() - start_time

            self.used_range = build_range_address(left=left, top=top, right=right, bottom=bottom)

        return self.used_range

    def get_sorted_merged_cells(self):
        """ 将合并单元格按照行列顺序排列。
        """
        if not hasattr(self, 'sorted_merged_cells'):
            self.sorted_merged_cells = list(sorted(self.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)))
        return self.sorted_merged_cells


inject_members(XlWorksheet, openpyxl.worksheet.worksheet.Worksheet)


class XlWorkbook(openpyxl.Workbook):

    def extract_summary(self, *, samples_num=5, limit_length=2500):
        """ 更新后的函数：提取整个Excel工作簿的摘要信息 """
        wb = self

        all_sheets_summary = []

        for ws in wb._sheets:  # 非数据表，也要遍历出来，所以使用了_sheets
            # 如果是标准工作表（Worksheet），使用现有的摘要提取机制
            if isinstance(ws, openpyxl.worksheet.worksheet.Worksheet):
                # 找到使用范围和表头范围
                used_range = ws.get_usedrange()
                if used_range:
                    header_range, data_range = split_header_and_data(ws, used_range)

                    # 提取表头结构
                    header_structure = extract_header_structure(ws, header_range)

                    filterRange = re.sub(r'\d+',
                                         lambda m: str(max(int(m.group()) - 1, 1)),
                                         data_range, count=1)

                    summary = ({
                        "sheetName": ws.title,
                        "sheetType": "Worksheet",
                        "usedRange": used_range,
                        "headerRange": header_range,
                        "header": header_structure,
                        'dataRange': data_range,
                        'filterRange': filterRange,
                        'sortRange': filterRange,
                        'data': extract_field_summaries(ws, header_range, data_range, samples_num)
                    })

                    if not summary['data']:  # 如果没有数据，则大概率是数据透视表，是计算出来的，读取不到~
                        summary['sheetType'] = 'PivotTable'
                        del summary['data']
                else:
                    summary = ({
                        "sheetName": ws.title,
                        "sheetType": "DialogOrMacroSheet",
                        "usedRange": None,
                    })

            # 如果是其他类型的工作表，提供基础摘要
            else:
                summary = ({
                    "sheetName": ws.title,
                    "sheetType": ws.__class__.__name__  # 使用工作表的类名作为类型
                })

            all_sheets_summary.append(summary)

        workbook_summary = {
            "fileName": Path(self.path).name if self.path else None,
            "sheetNames": wb.sheetnames,
            "sheets": all_sheets_summary,
        }

        WorkbookSummary(workbook_summary).reduce_summarys(limit_length=limit_length)

        return workbook_summary

    def extract_summary2(self):
        """ 另一套按照单元格提取摘要的程序 """
        wb = self

        all_sheets_summary = []

        for ws in wb._sheets:  # 非数据表，也要遍历出来，所以使用了_sheets
            # 如果是标准工作表（Worksheet），使用现有的摘要提取机制
            if isinstance(ws, openpyxl.worksheet.worksheet.Worksheet):
                # 找到使用范围和表头范围
                used_range = ws.get_usedrange()
                if used_range:
                    raw_used_range = ws.get_raw_usedrange()
                    summary = ({
                        "sheetName": ws.title,
                        "sheetType": "Worksheet",
                        "rawUsedRange": raw_used_range,
                        "usedRange": used_range,
                        'cells': extract_cells_content(ws)
                    })

                    if not summary['cells']:  # 如果没有数据，则大概率是数据透视表，是计算出来的，读取不到~ 但是JSA等场景应该有办法获得
                        summary['sheetType'] = 'PivotTable'
                        del summary['cells']
                else:
                    summary = ({
                        "sheetName": ws.title,
                        "sheetType": "DialogOrMacroSheet",
                        "usedRange": None,
                    })

            # 如果是其他类型的工作表，提供基础摘要
            else:
                summary = ({
                    "sheetName": ws.title,
                    "sheetType": ws.__class__.__name__  # 使用工作表的类名作为类型
                })

            all_sheets_summary.append(summary)

        workbook_summary = {
            "fileName": Path(self.path).name if self.path else None,
            "sheetNames": wb.sheetnames,
            "sheets": all_sheets_summary,
        }

        return workbook_summary


inject_members(XlWorkbook, openpyxl.Workbook)


def __3_extract_summary():
    """ 提取表格摘要 """


def score_row(row):
    score = 0
    for cell in row:
        if cell.value is not None:
            if is_string_type(cell.value):
                score += 1  # Add positive score for string type
            else:
                score -= 1  # Subtract score for non-string type

            # 检查填充颜色和边框，为得分增加0.5分
            # if cell.fill.bgColor.rgb != '00000000' or \
            #         (cell.border.left.style or cell.border.right.style or
            #          cell.border.top.style or cell.border.bottom.style):
            #     score += 0.5
    return score


def find_header_row(ws, used_range, max_rows_to_check=10):
    """ 找到工作表中的表头行 """
    range_details = parse_range_address(used_range)

    # 初始化得分列表
    row_scores = []

    # 只检查指定的最大行数
    rows_to_check = min(range_details['bottom'] - range_details['top'] + 1, max_rows_to_check)

    # 为每行评分
    for row in ws.iter_rows(min_row=range_details['top'], max_row=range_details['top'] + rows_to_check - 1,
                            min_col=range_details['left'], max_col=range_details['right']):
        row_scores.append(score_row(row))

    # 计算行与行之间分数变化的加权
    weighted_scores = []
    for i, score in enumerate(row_scores):
        b = score - row_scores[i + 1] if i < len(row_scores) - 1 else 0
        y = score + b
        weighted_scores.append(y)

    # 确定表头行的位置
    header_row = weighted_scores.index(max(weighted_scores)) + range_details['top']

    # 从used_range的起始行到找到的表头行都视为表头
    header_range = build_range_address(left=range_details['left'], top=range_details['top'],
                                       right=range_details['right'], bottom=header_row)
    return header_range


def split_header_and_data(ws, used_range, max_rows_to_check=50):
    """ 将工作表的used_range拆分为表头范围和数据范围 """
    header_range = find_header_row(ws, used_range, max_rows_to_check)
    header_details = parse_range_address(header_range)
    used_range_details = parse_range_address(used_range)

    # 数据范围是紧接着表头下面的部分，直到used_range的结束
    data_range = build_range_address(left=used_range_details['left'], top=header_details['bottom'] + 1,
                                     right=used_range_details['right'], bottom=used_range_details['bottom'])
    return header_range, data_range


def extract_header_structure(ws, header_range):
    """ 对输入的表头位置单元格，提取表头结构 """
    header_range_details = parse_range_address(header_range)

    header_structure = {}
    merged_addresses = set()

    # 处理合并的单元格
    for merged_range in ws.merged_cells.ranges:
        # 如果合并的单元格在提供的表头范围内
        if merged_range.bounds[1] <= header_range_details['bottom'] \
                and merged_range.bounds[3] >= header_range_details['top']:
            top_left_cell = ws.cell(row=merged_range.bounds[1], column=merged_range.bounds[0])
            address = build_range_address(left=merged_range.bounds[0], top=merged_range.bounds[1],
                                          right=merged_range.bounds[2], bottom=merged_range.bounds[3])
            header_structure[address] = top_left_cell.get_render_value()
            for row in range(merged_range.bounds[1], merged_range.bounds[3] + 1):
                for col in range(merged_range.bounds[0], merged_range.bounds[2] + 1):
                    merged_addresses.add((row, col))

    # 处理未合并的单元格
    for row in ws.iter_rows(min_row=header_range_details['top'], max_row=header_range_details['bottom'],
                            min_col=header_range_details['left'], max_col=header_range_details['right']):
        for cell in row:
            # 如果这个单元格的地址还没有被添加到结构中，并且它有一个值 （后记，没有值也添加下比较好）
            if (cell.row, cell.column) not in merged_addresses:
                header_structure[cell.coordinate] = cell.get_render_value()

    return header_structure


def determine_field_type_and_summary(ws, col, start_row, end_row, rows):
    """ 根据指定的列范围确定字段的摘要信息

    :param rows: 由外部传入要抽样的数据编号
    """
    # 1 需要全量读取数据，获知主要格式，和数值范围
    data = defaultdict(list)
    for i in range(start_row, end_row + 1):
        cell = ws.cell(i, col)
        k, v = cell.get_number_format(), cell.value
        data[k].append(v)

    data2 = sorted(data.items(), key=lambda item: len(item[1]), reverse=True)
    number_formats = [x[0] for x in data2]

    # 2 获得要展示的样本值
    sample_values = []
    for i in rows:
        cell = ws.cell(i, col)
        value = cell.get_render_value()
        if isinstance(value, str) and len(value) > 20:
            value = value[:17] + '...'
        sample_values.append(value)

    # 3 数值范围（只要判断主类型的数值范围就行了）
    numeric_range = None
    for x in data2:
        try:
            fmt, values = x
            values = [v for v in values if (v is not None and not isinstance(v, str))]
            numeric_range = [min(values), max(values)]
            numeric_range[0] = xl_render_value(numeric_range[0], fmt)
            numeric_range[1] = xl_render_value(numeric_range[1], fmt)
            break
        except (TypeError, ValueError) as e:
            pass

    summary = {
        "number_formats": number_formats,
        "numeric_range": numeric_range,
        "sample_values": sample_values,
    }

    return summary


def extract_cells_content(ws, usedrange=None):
    """ 提取一个工作表中的所有单元格内容

    1、找出所有合并单元格，按照我的那个顺序先排序。记为堆栈a（sorted_merged_cells_stack）。
        另外存一个已被使用的单元格集合b（初始空）（used_cells_set）。
    2、按照A1, A2, A3,..., B1, B2,B3等顺序遍历到一个个单元格c（cell）。
    3、对于每个c，先判断是否存在b中，如果存在b中，删除b中的c，并且跳过c的处理
    4、否则判断与a的最前面的元素a0是否相交，如果相交，则从堆栈a导出a0，把a0存储到结果数组cells。并且把a0中其他衍生单元格地址存入b。
    4、如果c不与a0相交，则c自身元素值写入cells。
    """
    # 1 预备
    sorted_merged_cells_stack = ws.get_sorted_merged_cells()[::-1]
    used_cells_set = set()
    if usedrange is None:
        usedrange = ws.get_usedrange()
    usedrange_bound = parse_range_address(usedrange)
    cells = {}  # 结果集合

    # 2 遍历所有单元格
    def get_val(cell):
        val = cell.value
        if val is None:
            return ''
        else:
            return cell.get_render_value()
            # 如果感觉这步很慢，可以换一种更简洁的形式；但是发现下面这种对时间的格式显示还是太不智能。
            # if not isinstance(val, (str, int, float, bool)):
            #     return val
            # else:
            #     return str(val)

    # 删除不在范围内的合并单元格
    sorted_merged_cells_stack = [
        mc for mc in sorted_merged_cells_stack
        if not (
            mc.min_row < usedrange_bound['top'] or
            mc.min_col < usedrange_bound['left'] or
            mc.max_col > usedrange_bound['right'] or
            mc.max_row > usedrange_bound['bottom']
        )
    ]


    for i in range(usedrange_bound['top'], usedrange_bound['bottom'] + 1):
        for j in range(usedrange_bound['left'], usedrange_bound['right'] + 1):
            # 2.1 合并单元格的衍生单元格，直接跳过
            if (i, j) in used_cells_set:
                used_cells_set.remove((i, j))
                continue
            cell = ws.cell(i, j)

            # 2.2 cell归属某组合并单元格（因为合并单元格排过序，所以只要很少的运算判断即可）
            # print(sorted_merged_cells_stack[-2])
            # print(sorted_merged_cells_stack[-2].min_row)
            # print(sorted_merged_cells_stack[-2].min_col)
            if (sorted_merged_cells_stack
                    and sorted_merged_cells_stack[-1].min_row == i
                    and sorted_merged_cells_stack[-1].min_col == j):
                rng = sorted_merged_cells_stack.pop()
                for rng_i in range(rng.min_row, rng.max_row + 1):
                    for rng_j in range(rng.min_col, rng.max_col + 1):
                        used_cells_set.add((rng_i, rng_j))
                cells[rng.coord] = get_val(cell)
                used_cells_set.remove((i, j))
                continue

            # 2.3 普通单元格
            # print(sorted_merged_cells_stack[-1].attributes())
            # print(sorted_merged_cells_stack[0].min_row,sorted_merged_cells_stack[0].max_row)
            cells[cell.coordinate] = get_val(cell)

    return cells


def extract_field_summaries(ws, header_range, data_range, samples_num=5):
    """ 提取所有字段的摘要信息

    :param samples_num: 要抽样的数据行数
    """
    # 1 数据范围信息
    header_details = parse_range_address(header_range)
    data_details = parse_range_address(data_range)
    start_row = header_details['bottom'] + 1
    end_row = data_details['bottom']

    # 2 提前决定好所有字段统一抽样的数据行号
    rows = list(range(header_details['bottom'] + 1, data_details['bottom'] + 1))
    if len(rows) > samples_num:
        rows = random.sample(rows, samples_num)
        rows.sort()

    # 3 提取所有字段的摘要信息
    field_summaries = {}
    for col in ws.iter_cols(min_col=header_details['left'], max_col=header_details['right']):
        header_cell = ws.cell(header_details['bottom'], col[0].column)
        if header_cell.celltype() != 1:  # todo 这里要改成不使用衍生单元格的场景
            # 注意，原本摘要这里用的是.value，后面改成了.coordinate。原本的遇到重名就会出一些问题了~
            field_summaries[header_cell.coordinate] = determine_field_type_and_summary(
                ws, header_cell.column, start_row, end_row, rows
            )

    return field_summaries


class WorkbookSummary:
    """ 工作薄摘要相关处理功能 """

    def __init__(self, data):
        self.data = data

    def reduce_summarys(self, limit_length=2500):
        """ 精简摘要

        :param limit_length: 参考限定的长度，折算后大概是限定1500token
        """
        sheets = self.data['sheets']

        if limit_length == -1:  # -1表示不限制长度
            return sheets

        # 1 不用改变
        text1 = json.dumps(sheets, ensure_ascii=False)
        length1 = len(text1)
        if length1 < limit_length:
            return sheets

        # 2 每个字段都去掉一个样本
        for sheet in sheets:
            if 'data' in sheet:
                st_data = sheet['data']
                for col_header, col in st_data.items():
                    st_data[col_header]['sample_values'] = col['sample_values'][:-1]

        text2 = json.dumps(sheets, ensure_ascii=False)
        length2 = len(text2)
        bias = length1 - length2
        # dprint(length1, length2, bias)  # 调试用
        if length2 <= limit_length:
            return sheets

        # 3 算出理论上要去掉几个样本，才能控制到理想长度
        n = math.ceil(safe_div(length1 - limit_length, bias))
        m = 5 - n
        if m >= 0:
            for sheet in sheets:
                if 'data' in sheet:
                    st_data = sheet['data']
                    for col_header, col in st_data.items():
                        if m > 0:
                            st_data[col_header]['sample_values'] = col['sample_values'][:m]
                        elif m == 0:
                            del st_data[col_header]['sample_values']
            return sheets

        # 4 如果m<0，可能靠上述删除还不够。这应该是不可能发生的事情，但还是处理下。
        for sheet in sheets:
            if 'data' in sheet:
                del sheet['data']
            # 这个删除后可能还是太长的话，表示sheet太多了，需要删掉一些sheet

        # 如果删除所有的数据后仍然超过限制，那么删除一些表格
        while len(json.dumps(sheets, ensure_ascii=False)) > limit_length:
            sheets.pop()

        self.data['sheets'] = sheets
        return sheets

    def random_filename(self):
        self.data['fileName'] = str(random.randint(0, 2000)) + '.xlsx'

    def choice_samples(self, samples_num=5):
        """ 限定最多抽取几个样本
        """
        data = self.data
        for sheet in data['sheets']:
            if 'data' in sheet:
                # 预设好要抽哪些行
                n = min([len(v['sample_values']) for k, v in sheet['data'].items()])
                rows = list(range(n))
                if len(rows) > samples_num:
                    rows = random.sample(rows, samples_num)
                    rows.sort()
                # 抽取样本
                for col_name, col_data in sheet['data'].items():
                    col_data['sample_values'] = [col_data['sample_values'][i] for i in rows]

    def random_delete(self):
        """ 随机删除一些字段 """
        data = self.data
        for sheet in data['sheets']:
            # 80%概率删除data
            if 'data' in sheet and random.random() < 0.8:
                del sheet['data']

            # filterRange和sortRange有80%概率随机删除一个
            if 'filterRange' in sheet and 'sortRange' in sheet:
                if random.random() < 0.5:
                    del sheet['filterRange']
                else:
                    del sheet['sortRange']

            # usedRange、headRange、dataRange，有20%概率随机删掉一个
            if 'usedRange' in sheet and 'headRange' in sheet and 'dataRange' in sheet:
                r = random.random()
                if r < 1 / 3:
                    del sheet['usedRange']
                elif r < 2 / 3:
                    del sheet['headRange']
                else:
                    del sheet['dataRange']

            # hearder有50%概率随机打乱
            if random.random() < 0.5:
                sheet['header'] = shuffle_dict_keys(sheet['header'])

        # 一半的概率删掉文件名
        if random.random() < 0.5:
            del data['fileName']

        # 如果只有一个sheet，还会进一步精简
        if len(data['sheets']) == 1:
            if random.random() < 0.5:
                data = data['sheets'][0]

            for name in ['sheetName', 'sheetType']:
                if random.random() < 0.5 and name in data:
                    del data[name]

    def to_str(self):
        return json.dumps(self.data, ensure_ascii=False)


def extract_workbook_summary(file_path, mode=0,
                             samples_num=5, limit_length=2500, ignore_errors=False):
    """ 更新后的函数：提取整个Excel工作簿的摘要信息

    :param mode:
        -1，提取全量摘要（详细信息，全部样本）
        0, 标准的提取摘要（详细信息，随机抽5个样本）
        1，精简摘要，在保留逻辑完整性的前提下，随机的修改一些摘要的结构内容
    """
    try:
        wb: XlWorkbook = openpyxl.load_workbook(file_path)
    except Exception as e:
        if ignore_errors:
            return {}
        else:
            raise e

    if mode == -1:
        res = wb.extract_summary(samples_num=1000, limit_length=-1)
        res['fileName'] = Path(file_path).name
    elif mode == 0:
        res = wb.extract_summary(samples_num=samples_num, limit_length=limit_length)
        res['fileName'] = Path(file_path).name

    elif mode == 1:
        res = wb.extract_summary(samples_num=samples_num)

        wb_summary = WorkbookSummary(res)
        wb_summary.random_filename()
        wb_summary.random_delete()
        wb_summary.reduce_summarys(limit_length=limit_length)

        res = wb_summary.data
    else:
        raise ValueError('mode参数值不正确')

    return res


def extract_workbook_summary2(file_path, *,
                              keep_links=False,
                              keep_vba=False,
                              mode=0,
                              return_mode=0,
                              **kwargs):
    """
    :param keep_links: 是否保留外部表格链接数据。如果保留，打开好像会有点问题。
    :param mode:
        0，最原始的summary3摘要
        1，添加当前工作表、单元格位置的信息
    :param kwargs: 捕捉其他参数，主要是向下兼容，其实现在并没有用

    注意这里没有提供read_only可选参数，是因为read_only=True模式在我这里是运行不了的。
    """

    # 1 读取文件wb
    def read_file_by_type():
        nonlocal load_time
        suffix = file_path.suffix.lower()
        start_time = time.time()
        if suffix in ('.xlsx', '.xlsm'):
            wb = openpyxl.load_workbook(file_path,
                                        keep_links=keep_links,
                                        keep_vba=keep_vba)
        elif suffix == '.xls':
            wb = convert_xls_to_xlsx(file_path)
        elif suffix == '.csv':
            wb = convert_csv_to_xlsx(file_path)
        else:
            return None
        load_time = time.time() - start_time
        return wb

    load_time = -1
    file_path = Path(file_path)
    res = {}
    res['fileName'] = file_path.name
    wb = read_file_by_type()
    if wb is None:  # 不支持的文件类型，不报错，只是返回最基本的文件名信息
        if return_mode == 1:
            return res, load_time
        else:
            return res

    # 2 提取摘要
    summary2 = wb.extract_summary2()
    DictTool.ior(res, summary2)
    if mode == 1:
        ws = wb.active
        res['ActiveSheet'] = ws.title
        if hasattr(ws, 'selected_cell'):
            res['Selection'] = ws.selected_cell

    # res = convert_to_json_compatible(res)

    if return_mode == 1:
        return res, load_time
    else:
        return res


def update_raw_summary2(data):
    # 1 中文率
    if 'chineseContentRatio' not in data:
        texts = [data['fileName']]  # 文件名和表格名都要加上
        texts += [x for x in data['sheetNames']]

        texts += [v for sheet in data['sheets'] for v in sheet.get('cells', {}).values() if v]
        all_text = ''.join(map(str, texts))
        data['chineseContentRatio'] = round(calc_chinese_ratio(all_text), 4)

    # 2 非空单元格率
    if 'nonEmptyCellRatio' not in data:
        content_area, total_area = 0, 0
        for sheet in data['sheets']:
            for addr, value in sheet.get('cells', {}).items():
                area = get_addr_area(addr)
                if value != '':
                    content_area += area
                total_area += area
        data['nonEmptyCellRatio'] = round(safe_div(content_area, total_area), 4)

    # 3 判断键值顺序
    keys = list(data.keys())
    ref_keys = ['fileName', 'chineseContentRatio', 'nonEmptyCellRatio', 'sheetNames', 'sheets']
    if keys != ref_keys:
        data = {k: data[k] for k in ref_keys if k in data}

    return data


def extract_workbook_summary2plus(file_path, **kwargs):
    """ 增加了全局ratio的计算 """
    # 1 主体摘要
    data = extract_workbook_summary2(file_path, **kwargs)
    if not data:
        return data

    # 2 增加一些特征计算
    # todo 后续估计要改成按table的颗粒度统计以下特征

    data = update_raw_summary2(data)
    return data


class WorkbookSummary3:
    """ 计算summary3及衍生版本需要的一些功能组件 """

    @classmethod
    def reduce1_delete_empty_cell(cls, summary3):
        """ 删除空单元格 """
        for sheet in summary3['sheets']:
            new_cells = {}
            for addr, val in sheet['cells'].items():
                if val != '':
                    new_cells[addr] = val
            sheet['cells'] = new_cells

    @classmethod
    def reduce2_truncate_overlong_cells(cls, summary3, summary_limit_len, *, cur_summary_len=None):
        """ 截断过长的单元格内容以满足表格摘要总长度限制。

        此算法旨在处理单个或多个单元格内容过长时，整个表格摘要总长度超过预设限制的问题。在保留尽可能多的有用信息的同时，智能地缩减内容长度。

        :param dict summary3: 表格摘要数据，包含多个sheet及其单元格内容
        :param int summary_limit_len: 表格摘要的最大长度限制
        :param int cur_summary_len: 当前表格摘要的长度，如果不提供，则会计算
        :return int: 调整后的表格摘要长度

        算法执行流程：
        1. 计算基准单元格长度和当前摘要长度超出部分（delta_length）。
        2. 筛选出超过基准长度的单元格并按长度降序排序。
        3. 逐个尝试截断单元格直到总长度满足要求或处理完所有单元格。
        """

        # 如果未提供当前摘要长度，则计算之
        if cur_summary_len is None:
            cur_summary_len = len(json.dumps(summary3, ensure_ascii=False))

        # 1. 计算基准单元格长度
        total_cells_num = sum(len(st['cells']) + 5 for st in summary3['sheets'])
        base_cell_length = int(-60 * math.log(total_cells_num, 10) + 260)  # 从200趋近到20
        base_cell_length = min(int(summary_limit_len * 0.05), base_cell_length)
        base_cell_length = max(int(summary_limit_len * 0.005), base_cell_length)

        # 2. 预提取全部单元格数据信息并计算需要减少的长度
        delta_length = cur_summary_len - summary_limit_len  # 计算需要减少的长度
        overlong_cells = [(sheet, addr, val, len(val)) for sheet in summary3['sheets']
                          for addr, val in sheet['cells'].items() if
                          isinstance(val, str) and len(val) > base_cell_length]
        overlong_cells.sort(key=lambda x: -x[3])  # 按长度降序排序

        # 3. 逐个尝试截断单元格直到满足长度要求
        possible_reduction = 0
        for i, (_, _, _, length) in enumerate(overlong_cells):
            next_len = overlong_cells[i + 1][3] if i + 1 < len(overlong_cells) else base_cell_length
            possible_reduction += (length - next_len) * (i + 1)
            if possible_reduction >= delta_length or i == len(overlong_cells) - 1:
                for j in range(i + 1):
                    sheet, addr, val, _ = overlong_cells[j]
                    sheet['cells'][addr] = val[:next_len - 3] + '...'  # 更新单元格内容
                break

        return cur_summary_len - possible_reduction

    @classmethod
    def reduce3_fold_rows(cls, summary3, summary_limit_len, *, cur_summary_len=None):
        if cur_summary_len is None:
            cur_summary_len = len(json.dumps(summary3, ensure_ascii=False))

        # 每个sheet本身其他摘要，按照5个单元格估算
        total_cells_num = sum([(len(st['cells']) + 5) for st in summary3['sheets']])
        avg_cell_len = cur_summary_len / total_cells_num
        # 目标删除单元格数量，向上取整
        target_reduce_cells_num = int((cur_summary_len - summary_limit_len) / avg_cell_len + 0.5)

        # 相同区域，头尾至少要留2行，然后考虑压缩量，一般一块range至少要10行同构数据，进行中间至少6行的压缩描述才有意义
        # 考虑重要性，应该是从末尾表格，末尾数据往前检索压缩，直到压缩量满足要求

        for sheet in reversed(summary3['sheets']):
            cells = sheet['cells']
            # 1 对单元格，先按行分组
            last_line_id = -1
            row_groups = []
            for addr, val in cells.items():
                m = re.search(r'\d+', addr)
                if not m:  # 应该都一定能找到的，这个判断是为了防止报错
                    continue
                line_id = int(m.group())

                val_type_tag = '@' if isinstance(val, str) else '#'
                cell_tag = re.sub(r'\d+', '', addr) + val_type_tag

                if line_id == last_line_id:
                    row_groups[-1].append([addr, cell_tag])
                else:
                    row_groups.append([[addr, cell_tag]])
                last_line_id = line_id

            # 2 算出每一行的row_tag，并按照row_tag再分组
            last_row_tag = ''
            rows_groups = []
            for row in row_groups:
                row_tag = ''.join([cell_tag for _, cell_tag in row])
                if row_tag == last_row_tag:
                    rows_groups[-1].append(row)
                else:
                    rows_groups.append([row])
                last_row_tag = row_tag

            # 3 开始压缩
            def extract_cells_from_rows(rows):
                for row in rows:
                    for addr, _ in row:
                        new_cells[addr] = cells[addr]

            new_cells = {}
            for rows in rows_groups:
                if len(rows) < 10:
                    extract_cells_from_rows(rows)
                else:  # 压缩中间的数据
                    # 如果评估到最终摘要可能太小，要收敛下删除的范围
                    n, m = len(rows), len(rows[0])
                    target_n = int(target_reduce_cells_num / m + 0.5)  # 本来应该删除多少行才行
                    cur_n = n - 4 if target_n > n - 4 else target_n  # 实际删除多少行
                    left_n = n - cur_n  # 剩余多少行
                    b = left_n // 2
                    a = left_n - b

                    extract_cells_from_rows(rows[:a])
                    addr = combine_addresses(rows[a][0][0], rows[-b - 1][-1][0])
                    # new_cells[addr] = '这块区域的内容跟前面几行、后面几行的内容结构是一致的，省略显示'
                    new_cells[addr] = '...'
                    extract_cells_from_rows(rows[-b:])

                    target_reduce_cells_num -= cur_n * m
                    if target_reduce_cells_num <= 0:
                        break

            sheet['cells'] = new_cells

    @classmethod
    def reduce4_truncate_cells(cls, y, summary_limit_len, *, cur_summary_len=None):
        if cur_summary_len is None:
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))

        # 1 预计要删除单元格数
        sheet_cells_num = [len(st['cells']) for st in y['sheets']]
        # 每个sheet本身其他摘要，按照5个单元格估算
        total_cells_num = sum(sheet_cells_num) + len(sheet_cells_num) * 5
        avg_cell_len = cur_summary_len / total_cells_num
        # 目标删除单元格数量，向上取整
        target_reduce_cells_num = int((cur_summary_len - summary_limit_len) / avg_cell_len + 0.5)

        # 2 所有的单元格如果都不够删，那就先把所有cells删了再说
        if total_cells_num < target_reduce_cells_num:
            for st in y['sheets']:
                st['cells'] = {}
            return len(json.dumps(y, ensure_ascii=False))

        # 3 否则每张表按照比例删单元格，只保留前面部分的单元格
        left_rate = 1 - target_reduce_cells_num / total_cells_num
        while True:
            for i, st in enumerate(y['sheets']):
                st['cells'] = dict(islice(st['cells'].items(), int(left_rate * sheet_cells_num[i])))
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))
            if cur_summary_len <= summary_limit_len:
                return cur_summary_len
            if left_rate * total_cells_num < 1:
                break
            else:  # 缩小保留比例，再试
                left_rate *= 0.8

        return cur_summary_len

    @classmethod
    def reduce5_truncate_sheets(cls, y, summary_limit_len, *, cur_summary_len=None):
        """ 计算平均每张表的长度，保留前面部分的表格 """
        if cur_summary_len is None:
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))

        n = len(y['sheets'])
        avg_sheet_len = cur_summary_len / n
        target_reduce_sheet_num = int((cur_summary_len - summary_limit_len) / avg_sheet_len + 0.5)
        y['sheets'] = y['sheets'][:n - target_reduce_sheet_num]

        while y['sheets']:
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))
            if cur_summary_len <= summary_limit_len:
                return cur_summary_len
            y['sheets'] = y['sheets'][:-1]  # 依次尝试删除最后一张表格的详细信息

    @classmethod
    def summary2_to_summary3(cls, summary2, summary_limit_len=4000):
        """ 将summary2转换成summary3 """

        def reduce_step_by_step(y):
            mode_tags = [
                'Delete empty cell',
                'Omit the longer content and replace it with "..."',
                'Omit lines with the same structure',
                'Omit later lines',
                'Omit later sheets'
            ]

            # 0 摘要本来就不大
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))
            if cur_summary_len <= summary_limit_len:
                return y

            # 1 删除空单元格
            cls.reduce1_delete_empty_cell(y)
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:1])
                return y

            # 2 单个单元格内容过长的，省略显示
            cur_summary_len = cls.reduce2_truncate_overlong_cells(y, summary_limit_len, cur_summary_len=cur_summary_len)
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:2])
                return y

            # 3 同构数据，省略显示（有大量相同行数据，折叠省略表达）
            cls.reduce3_fold_rows(y, summary_limit_len, cur_summary_len=cur_summary_len)
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:3])
                return y

            # 4 每张表都按比例删除后面部分的单元格
            cur_summary_len = cls.reduce4_truncate_cells(y, summary_limit_len, cur_summary_len=cur_summary_len)
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:4])
                return y

            # 5 从后往前删每张表格的详细信息
            cls.reduce5_truncate_sheets(y, summary_limit_len, cur_summary_len=cur_summary_len)
            y['mode'] = ', '.join(mode_tags[:5])
            return y

        x = summary2
        y = {
            'fileName': x['fileName'],
            'sheetNames': x['sheetNames'],
            'sheets': x['sheets'],
            'mode': 'Complete information',
        }

        # 处理前确保下cells字段存在，避免后续很多处理过程要特判
        for st in y['sheets']:
            if 'cells' not in st:
                st['cells'] = {}

        y = reduce_step_by_step(y)

        # 但是最后结果还是去掉空cells
        for st in y['sheets']:
            if not st['cells']:
                del st['cells']

        return y

    @classmethod
    def reduce4b(cls, y, summary_limit_len, *, cur_summary_len=None, active_sheet_weight=0.5):
        """
        :param active_sheet_weight: 当前活动表格被删除的权重，0.5表示按比例被删除的量只有其他表格的一半
        """
        if cur_summary_len is None:
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))

        active_sheet = y['ActiveSheet']

        # 1 预计要删除单元格数
        sheet_cells_num = [len(st['cells']) for st in y['sheets']]
        # 每个sheet本身其他摘要，按照5个单元格估算
        total_cells_num = sum(sheet_cells_num) + len(sheet_cells_num) * 5
        avg_cell_len = cur_summary_len / total_cells_num
        # 目标删除单元格数量，向上取整
        target_reduce_cells_num = int((cur_summary_len - summary_limit_len) / avg_cell_len + 0.5)

        # 2 对当前活动表格，会减小删除权重
        # 标记当前活动表格的单元格数
        active_sheet_index = [i for i, st in enumerate(y['sheets']) if st['sheetName'] == active_sheet][0]
        active_cells_num = sheet_cells_num[active_sheet_index]

        # 计算权重系数
        w = active_sheet_weight  # 当前激活表的权重系数
        m = active_cells_num
        n = total_cells_num
        r = target_reduce_cells_num / n

        # 计算非活动表格的额外权重系数
        w2 = 1 + m * (1 - w) / (n - m)

        # 3 所有的单元格如果都不够删，那就先把所有cells删了再说
        if total_cells_num < target_reduce_cells_num:
            for st in y['sheets']:
                st['cells'] = {}
            return len(json.dumps(y, ensure_ascii=False))

        # 4 否则每张表按照比例删单元格，只保留前面部分的单元格
        left_rate = 1 - r  # 原始保留比例
        while True:
            for i, st in enumerate(y['sheets']):
                if i == active_sheet_index:
                    # 当前激活的sheet保留更多单元格
                    st['cells'] = dict(islice(st['cells'].items(), int(left_rate * w * sheet_cells_num[i])))
                else:
                    # 其他sheet按照w2权重删除单元格
                    st['cells'] = dict(islice(st['cells'].items(), int(left_rate * w2 * sheet_cells_num[i])))
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))
            if cur_summary_len <= summary_limit_len:
                return cur_summary_len
            if left_rate * total_cells_num < 1:
                break
            else:
                left_rate *= 0.8  # 缩小保留比例，再试

        return cur_summary_len

    @classmethod
    def reduce5b(cls, y, summary_limit_len, *, cur_summary_len=None):
        """ 计算平均每张表的长度，保留前面部分的表格 """
        if cur_summary_len is None:
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))

        n = len(y['sheets'])
        active_sheet_name = y['ActiveSheet']

        avg_sheet_len = cur_summary_len / n
        # target_reduce_sheet_num = int((cur_summary_len - summary_limit_len) / avg_sheet_len + 0.5)
        # y['sheets'] = y['sheets'][:n - target_reduce_sheet_num]

        while y['sheets']:
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))
            if cur_summary_len <= summary_limit_len:
                return cur_summary_len

            # 如果最后一张表格是激活的表格，尝试删除前一张
            if y['sheets'][-1]['sheetName'] == active_sheet_name:
                if len(y['sheets']) > 1:
                    y['sheets'] = y['sheets'][:-2] + [y['sheets'][-1]]
                else:
                    y['sheets'] = []
            else:
                y['sheets'] = y['sheets'][:-1]  # 删除最后一张表格的详细信息

        return cur_summary_len

    @classmethod
    def summary2_to_summary3b(cls, summary2, summary_limit_len=4000):
        """ 将summary2转换成summary3 """

        def reduce_step_by_step(y):
            mode_tags = [
                'Delete empty cell',
                'Omit the longer content and replace it with...',
                'Omit lines with the same structure',
                'Omit later lines',
                'Omit later sheets'
            ]

            # 0 摘要本来就不大
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))
            if cur_summary_len <= summary_limit_len:
                return y

            # 1 删除空单元格
            cls.reduce1_delete_empty_cell(y)
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:1])
                return y

            # 2 单个单元格内容过长的，省略显示
            cur_summary_len = cls.reduce2_truncate_overlong_cells(y, summary_limit_len, cur_summary_len=cur_summary_len)
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:2])
                return y

            # 3 同构数据，省略显示（有大量相同行数据，折叠省略表达）
            cls.reduce3_fold_rows(y, summary_limit_len, cur_summary_len=cur_summary_len)
            cur_summary_len = len(json.dumps(y, ensure_ascii=False))
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:3])
                return y

            # 4 每张表都按比例删除后面部分的单元格
            cur_summary_len = cls.reduce4b(y, summary_limit_len, cur_summary_len=cur_summary_len)
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:4])
                return y

            # 5 从后往前删每张表格的详细信息
            cls.reduce5b(y, summary_limit_len, cur_summary_len=cur_summary_len)
            y['mode'] = ', '.join(mode_tags[:5])
            return y

        x = summary2
        y = {
            'fileName': x['fileName'],
            'sheetNames': x['sheetNames'],
            'sheets': x['sheets'],
            'mode': 'Complete information',
            'ActiveSheet': x['ActiveSheet'],  # 当期激活的工作表
        }
        if 'Selection' in x:
            # 最多截取250个字符。（一般情况下这个很小的，只是在很极端情况，比如离散选中了非常多区域，这个可能就会太长
            y['Selection'] = x['Selection'][:250]

        # 处理前确保下cells字段存在，避免后续很多处理过程要特判
        for st in y['sheets']:
            if 'cells' not in st:
                st['cells'] = {}

        y = reduce_step_by_step(y)

        # 但是最后结果还是去掉空cells
        for st in y['sheets']:
            if not st['cells']:
                del st['cells']

        return y


def extract_workbook_summary3(file_path, summary_limit_len=4000, **kwargs):
    """ 增加了全局ratio的计算 """
    data = extract_workbook_summary2(file_path, **kwargs)
    if not data:
        return data
    data = WorkbookSummary3.summary2_to_summary3(data, summary_limit_len)
    return data


def extract_workbook_summary3b(file_path,
                               summary_limit_len=4096,
                               timeout_seconds=10,
                               return_mode=0,
                               debug=False,
                               **kwargs):
    """

    :param summary_limit_len: 摘要长度限制
    :param timeout_seconds: 超时限制
    :param return_mode: 返回模式，0表示只返回摘要，1表示返回摘要和耗时
    :param kwargs: 其他是summary2读取文件的时候的参数，其实都不太关键，一般不用特地设置
    """
    res = {}
    res['fileName'] = Path(file_path).name
    load_time = summary2_time = summary3_time = -1

    try:
        # with Timeout(timeout_seconds):
        start_time = time.time()
        res, load_time = extract_workbook_summary2(file_path, mode=1, return_mode=1, **kwargs)
        # res = convert_to_json_compatible(res)
        summary2_time = time.time() - start_time - load_time
        start_time = time.time()
        res = WorkbookSummary3.summary2_to_summary3b(res, summary_limit_len)
        summary3_time = time.time() - start_time
    except TimeoutError as e:
        if debug:
            raise e
        res['error'] = f'超时，未完成摘要提取：{timeout_seconds}秒'
    except Exception as e:
        if debug:
            raise e
        res['error'] = f'提取摘要时发生错误：{format_exception(e, 2)}'

    if return_mode == 1:
        return res, {'load_time': human_readable_number(load_time),
                     'summary2_time': human_readable_number(summary2_time),
                     'summary3_time': human_readable_number(summary3_time)}
    return res


def test_speed():
    from tqdm import tqdm

    data, elaps = extract_workbook_summary3b(r'05成绩表.xlsx', return_mode=1)
    print(json.dumps(data, indent=2, ensure_ascii=False))
    print(elaps)

    for i in tqdm(range(1000000)):
        data = extract_workbook_summary3b(r'05成绩表.xlsx')

if __name__ == "__main__":
    os.chdir(os.environ.get("FILE_SYSTEM_PATH"))
    file_path = './excel_data/BikeBuyers_Data.xlsx'
    print(extract_workbook_summary3b(file_path))